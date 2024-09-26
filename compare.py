import re
from openpyxl import load_workbook

# 1. 从 Excel 文件中读取数据并处理
def read_excel_data(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active  # 或者指定工作表名称，如 wb['Sheet1']

    excel_data = []
    # 遍历第 D 列，从第二行开始
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        cell_value = row[0].value
        if cell_value is not None:
            # 丢掉前两个字符
            modified_value = str(cell_value)[2:]
            excel_data.append(modified_value)
    return excel_data

# 2. 从 txt 文件中提取数据并处理
def read_txt_data(txt_file):
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()

    txt_data = []
    # 使用正则表达式提取以 A99A 开头，0D0A 结尾的数据段
    pattern = r'A99A.*?0D0A'
    matches = re.findall(pattern, content)
    for match in matches:
        # 丢掉前两个字符
        modified_match = match[2:]
        txt_data.append(modified_match)
    return txt_data

# 3. 比较两个数据列表
def compare_data(excel_data, txt_data):
    min_length = min(len(excel_data), len(txt_data))
    for i in range(min_length):
        excel_item = excel_data[i]
        txt_item = txt_data[i]
        if excel_item == txt_item:
            print(f"第 {i+1} 项匹配：\nExcel 数据：{excel_item}\nTXT 数据：{txt_item}\n")
        else:
            print(f"第 {i+1} 项不匹配：\nExcel 数据：{excel_item}\nTXT 数据：{txt_item}\n")

# 主函数
def main():
    excel_file = 'data/11_tcpserver.xlsx'  # 替换为您的 Excel 文件名
    txt_file = 'data/11_log解析.txt'       # 替换为您的 txt 文件名

    # excel_file = 'data/11_tcpserver.xlsx'  # 替换为您的 Excel 文件名
    # txt_file = 'data/11_log解析.txt'  # 替换为您的 txt 文件名

    excel_data = read_excel_data(excel_file)
    txt_data = read_txt_data(txt_file)

    compare_data(excel_data, txt_data)

if __name__ == "__main__":
    main()
